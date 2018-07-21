import scrapy
from scrapy.crawler import CrawlerProcess
import scraperwiki



# -*- coding: utf-8 -*-
import scrapy
from openpyxl import load_workbook, Workbook
from copy import copy
import logging
import urllib
import os
from scrapy.crawler import CrawlerProcess
import json
import requests
import datetime

logging.basicConfig(level=logging.DEBUG)
base_url = 'https://www.fernsehserien.de/{}/sendetermine'

fetch_short_header = {
    "Origin": "https://www.fernsehserien.de",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
    "Content-type": "application/x-www-form-urlencoded",
    "Accept": "*/*",
    "Referer": "https://www.fernsehserien.de/malcolm-mittendrin",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,ro;q=0.6"
}


smart_selector = datetime.datetime.now().year - 5
data_header = ['Datum', 'Source: Tunesat', 'Sendejahr', 'Spot', 'TV', 'Sprecher', 'Other', 'Sendezeit Film / TV Minuten', 'Kommentar', 'Kanal', 'Land', 'Sendung', 'Episode / Details', 'Performer 1', 'Performer 2', 'Performer 3', 'Performer 4', 'Notiz', 'Quelle (Datum/Zustellung Brief)', 'Filepath', 'Album', 'Title TS', 'Title Name BMI', 'Show Data', 'Date', 'Time', 'Channel', 'Episode Number', 'Season', 'Episode', 'Episode Title']


print(os.getcwd())
filepath = 'res/rohdaten1.xlsx'

class FernsehseDeSpider(scrapy.Spider):
    name = 'fernsehse_de'
    allowed_domains = ['fernsehse']
    start_urls = ['http://fernsehse/']

    def copy_headers(self,write_sheet):
        logging.info('output_{} has copied header'.format(self.xlsx_counter))
        h = ['Show Data', 'Date', 'Time', 'Channel', 'Episode Number', 'Season', 'Episode', 'Episode Title']
        for ridx, row in enumerate(self.ws.iter_rows(max_col=self.ws.max_column, min_row=1, max_row=1), 1):
            for idx, cell in enumerate(row, 1):
                write_sheet.cell(ridx, idx, cell.value)
                if cell.has_style:
                    write_sheet.cell(ridx, idx).font = copy(cell.font)
                    write_sheet.cell(ridx, idx).border = copy(cell.border)
                    write_sheet.cell(ridx, idx).fill = copy(cell.fill)
                    write_sheet.cell(ridx, idx).number_format = copy(cell.number_format)
                    write_sheet.cell(ridx, idx).protection = copy(cell.protection)
                    write_sheet.cell(ridx, idx).alignment = copy(cell.alignment)

            for nextidx, _ in enumerate(h, idx + 1):
                write_sheet.cell(1, nextidx, _)

    def __init__(self):
        logging.info('directory {}'.format(os.getcwd()))
        self.wb = load_workbook(filepath)
        self.ws = self.wb.active
        self.xlsx_counter = 1

        self.document = None
        self.write_sheet = None

        self.rows= []

    def start_requests(self):
        # self.wb = load_workbook('../res/rohdaten1.xlsx')
        # self.ws = self.wb.active
        self.dublicates = []
        self.old_name = ''
        self.header = None

        # for url in urls:
        #     yield scrapy.Request(url=url, callback=self.parse)

        for ridx, row in enumerate(self.ws.iter_rows(max_col=self.ws.max_column, min_row=1, max_row=self.ws.max_row), 1):
            title = self.ws['L{}'.format(ridx)].value

            if title and title not in self.dublicates:
                if not self.document:
                    self.document = Workbook()
                    self.write_sheet = self.document.active
                    rows = []
                    self.copy_headers(self.write_sheet)
                    if ridx == 1:
                        continue

                self.header = copy([_.value for _ in row])
                self.header.append(None)

                if title != self.old_name:
                    l = [_.value for _ in row]
                    l.append(None)
                    self.rows.append(l)

                self.old_name = title

                url = 'https://www.fernsehserien.de/fastsearch'

                search_header = {
                    "Host": "www.fernsehserien.de",
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:61.0) Gecko/20100101 Firefox/61.0",
                    "Accept": "*/*",
                    "Accept-Language": "en-GB,en;q=0.5",
                    "Accept-Encoding": "gzip, deflate, br",
                    "Referer": "https://www.fernsehserien.de/suche/{}".format(urllib.parse.quote(title)),
                    "Content-type": "application/x-www-form-urlencoded",
                    "Content-Length": "24",
                    "Connection": "keep-alive"
                }
                search_results = requests.post(url, data={"suchwort": title}, headers=search_header)
                search_results = json.loads(search_results.text)

                for idx, data in enumerate(search_results, 1):
                    if 3 < idx:
                        break
                    data_url = base_url.format(data['s'])
                    yield scrapy.Request(url=data_url, callback=self.prepose, headers=fetch_short_header)


    def one_page_parse(self,response):
        prev = None
        logging.info('parse called for {}'.format(response.url))
        # for row_data in response.css('table.sendetermine tbody'):
        # for idx, row_data in enumerate(response.xpath("//table[contains(@class, 'sendetermine')]/tbody"),1):
        for idx, row_data in enumerate(response.xpath(
                "//table[contains(@class, 'sendetermine')]/tbody[@style='cursor:pointer']|//table[contains(@class, 'sendetermine')]/tbody"),
                                       1):
            remote_dt = row_data.css('tr td.sendetermine-datum::text')
            if remote_dt:
                remote_dt = remote_dt.extract_first()
            else:
                prev = row_data.xpath('preceding-sibling::*')
                if prev:
                    prev = prev[-1]
                    remote_dt = remote_dt.css('tr td.sendetermine-datum::text')
                else:
                    remote_dt = 'N/A'

            # remote_time = row_data.css('tr td.no-smartphone2::text')
            # remote_time = row_data.css('td:nth-child(4) > span:nth-child(1)')
            remote_time = row_data.xpath("tr/td[contains(@class, 'no-smartphone2')]/span/text()")
            if remote_time:
                remote_time = remote_time.extract_first()
            elif prev:
                remote_time = prev.xpath("tr/td[contains(@class, 'no-smartphone2')]/span/text()").extract_first()
            else:
                remote_time = 'N/A'

            # channel = row_data.css('tr td.sendetermine-sender::text')
            channel = row_data.xpath("tr/td[contains(@class,'sendetermine-sender')]/descendant-or-self::*/text()")
            if channel:
                channel = ' '.join(channel.extract())
            elif prev:
                channel = prev.xpath(
                    "tr/td[contains(@class,'sendetermine-sender')]/descendant-or-self::*/text()").extract()
                channel = ' '.join(channel)
            else:
                channel = 'N/A'

            episode_number = row_data.css("tr:nth-child(1) > td:nth-child(7) > b:nth-child(1)")

            if episode_number:
                episode_number = episode_number.css("::text")
                episode_number = episode_number.extract_first() if episode_number else 'N/A'
            else:
                episode_number = 'N/A'

            season = row_data.xpath(
                'tr/td[@align="right" and @style="padding-right:0"]/span/text()|td[contains(@style,"padding-right:0;")]/b/text()')
            if season:
                season = season.extract_first()
            else:
                season = 'N/A'

            episode = row_data.xpath('tr[1]/td[10]/span/text()')
            if episode:
                episode = episode.extract_first()

            episode_title = row_data.xpath('tr[1]/td[12]/b/text()')
            if episode_title:
                episode_title = episode_title.extract_first()

            row = [*self.header, remote_dt, remote_time, channel, episode_number, season, episode, episode_title]

            # row = [remote_dt, remote_time, channel, episode_number, season, episode, episode_title]
            yield {k: v for k, v in zip(data_header, row)}
            logging.info('got {}'.format(row))

        next_page = response.css('a.fore:nth-child(1) ::attr("href")')
        if next_page:
            next_page = 'https://www.fernsehserien.de{}'.format(next_page.extract_first())
            yield response.follow(next_page, self.parse)

    def prepose(self, response):
        logging.info('called preparse')
        select_option = []
        for option in response.css('#sendetermine-select--jahr option'):
            opt = option.css('option::text').extract_first()
            value = option.xpath('@value').extract_first()
            if 'ab jetzt' not in opt and 'Chronik' not in opt:
                select_option.append((opt, value))

        available_options = [_ for _ in select_option if smart_selector >= int(_[0])]

        if available_options:
            # select_option.sort(key=(lambda _: int(_[0])))
            select_option = available_options[0][1]
            page = 'https://www.fernsehserien.de{}'.format(select_option)
            logging.info('selected from option {}'.format(page))
            yield scrapy.Request(url=page, callback=self.parse, headers=fetch_short_header)
        elif select_option and len(select_option)>1:
            select_option.sort(key=(lambda _: int(_[0])))
            select_option = select_option[0][1]
            page = 'https://www.fernsehserien.de{}'.format(select_option)
            logging.info('selected from option {}'.format(page))
            yield scrapy.Request(url=page, callback=self.parse, headers=fetch_short_header, dont_filter = True)
        else:
            res = dict.fromkeys(data_header, None)
            res2 = {k:v for k,v in zip(data_header, self.header)}
            res = {**res, **res2}
            res['Show Data'] = 'NO'
            scraperwiki.sqlite.save(unique_keys=[], data=res)

    def parse(self, response):
        prev = None
        logging.info('parse called for {}'.format(response.url))
        # for row_data in response.css('table.sendetermine tbody'):
        # for idx, row_data in enumerate(response.xpath("//table[contains(@class, 'sendetermine')]/tbody"),1):
        for idx, row_data in enumerate(response.xpath("//table[contains(@class, 'sendetermine')]/tbody[@style='cursor:pointer']|//table[contains(@class, 'sendetermine')]/tbody"),1):
            remote_dt = row_data.css('tr td.sendetermine-datum::text')
            if remote_dt:
                remote_dt = remote_dt.extract_first()
            else:
                prev = row_data.xpath('preceding-sibling::*')
                if prev:
                    prev = prev[-1]
                    remote_dt = remote_dt.css('tr td.sendetermine-datum::text')
                else:
                    remote_dt = 'N/A'

            # remote_time = row_data.css('tr td.no-smartphone2::text')
            # remote_time = row_data.css('td:nth-child(4) > span:nth-child(1)')
            remote_time = row_data.xpath("tr/td[contains(@class, 'no-smartphone2')]/span/text()")
            if remote_time:
                remote_time = remote_time.extract_first()
            elif prev:
                remote_time = prev.xpath("tr/td[contains(@class, 'no-smartphone2')]/span/text()").extract_first()
            else:
                remote_time = 'N/A'

            # channel = row_data.css('tr td.sendetermine-sender::text')
            channel = row_data.xpath("tr/td[contains(@class,'sendetermine-sender')]/descendant-or-self::*/text()")
            if channel:
                channel = ' '.join(channel.extract())
            elif prev:
                channel = prev.xpath("tr/td[contains(@class,'sendetermine-sender')]/descendant-or-self::*/text()").extract()
                channel = ' '.join(channel)
            else:
                channel = 'N/A'

            episode_number = row_data.css("tr:nth-child(1) > td:nth-child(7) > b:nth-child(1)")

            if episode_number:
                episode_number = episode_number.css("::text")
                episode_number = episode_number.extract_first() if episode_number else 'N/A'
            else:
                episode_number = 'N/A'

            season = row_data.xpath('tr/td[@align="right" and @style="padding-right:0"]/span/text()|td[contains(@style,"padding-right:0;")]/b/text()')
            if season:
                season = season.extract_first()
            else:
                season = 'N/A'

            episode = row_data.xpath('tr[1]/td[10]/span/text()')
            if episode:
                episode = episode.extract_first()

            episode_title = row_data.xpath('tr[1]/td[12]/b/text()')
            if episode_title:
                episode_title = episode_title.extract_first()

            row = [*self.header, remote_dt, remote_time, channel, episode_number, season, episode, episode_title]

            # row = [remote_dt, remote_time, channel, episode_number, season, episode, episode_title]

            store =  {k:v for k,v in zip(data_header, row)}
            scraperwiki.sqlite.save(unique_keys=[], data=store)
            logging.info('got {}'.format(row))

        next_page = response.css('a.fore:nth-child(1) ::attr("href")')
        if next_page:
            next_page = 'https://www.fernsehserien.de{}'.format(next_page.extract_first())
            yield response.follow(next_page, self.parse)



if __name__ == "__main__":
    process = CrawlerProcess({
        'USER_AGENT': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)'
    })


    process.start() # the script will block here until the crawling is finished



scraperwiki.sqlite.execute("DROP table data")
process = CrawlerProcess()
process = CrawlerProcess({
        'USER_AGENT': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)'
    })
process.crawl(FernsehseDeSpider)
process.start()